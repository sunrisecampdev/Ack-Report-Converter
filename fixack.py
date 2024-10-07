import openpyxl
from os import path
from pprint import pprint
from donor import *

class RevenueReport:
    def __init__(self, workbook, sheetRaw, sheetFormat, sheetMultiple, sheetSingle, sheetNothing, sheetOpen):
        self.workbook = workbook
        self.sheetRaw = sheetRaw
        self.sheetFormat = sheetFormat
        self.sheetMultiple = sheetMultiple
        self.sheetSingle = sheetSingle
        self.sheetNothing = sheetNothing
        self.sheetOpen = sheetOpen
        self.HEADER_ROW = self.getHeaderRowIndex()-1
        self.MAX_COL = self.sheetRaw.max_column+1
        self.MAX_ROW = self.sheetRaw.max_row
        self.colIndex = 1
        self.rowIndex = self.getHeaderRowIndex()+1
        self.headerDict = dict({
            "DONOR_ID" : {"name" : "Donor ID"},
            "SPOUSE_FIRST" : {"name" : "Household: Spouse First Name"},
            "SPOUSE_LAST" : {"name" : "Household: Spouse Last Name"},
            "PRIMARY_FIRST" : {"name": "Household: Primary First Name"},
            "PRIMARY_LAST" : {"name" : "Household: Primary Last Name"},
            "COMPANY" : {"name" : "Company Name"},
            "EVENT_STATUS" : {"name": "Event Status"},
            "ATTENDEES" : {"name": "Attendees"},
            "DATE" : {"name": "Date"},
            "SALE_TYPE" : {"name" : "Type"},
            "AMOUNT" : {"name" : "Amount"},
            "CONNECTION" : {"name" : "Attribute"},
            "FIRST_NAME" : {"name" : "First Name"},
            "LAST_NAME" : {"name" : "Last Name"},
            "DONOR" : {"name" : "Donor"},
            "PAID" : {"name" : "Paid"},
            "ADDRESS1" : {"name" : "Address 1"},
            "ADDRESS2" : {"name" : "Address 2"},
            "CITY" : {"name" : "City"},
            "STATE" : {"name" : "State"},
            "ZIP" : {"name" : "Zip Code"},
            "EMAIL" : {"name" : "Email"},
            "ALTEMAIL" : {"name" : "Alt. Email"}
        })
        self.headerOrder = ("DONOR_ID",
                            "FIRST_NAME",
                            "LAST_NAME",
                            "COMPANY",
                            "ADDRESS1",
                            "ADDRESS2",
                            "CITY",
                            "STATE",
                            "ZIP",
                            "EVENT_STATUS",
                            "SALE_TYPE",
                            "AMOUNT",
                            "EMAIL",
                            "ALTEMAIL",
                            "DATE"
                            )
        self.headerIndexMap = dict()
        self.sheetFormatRowIndex = 2

        self.mapColIndices()
        self.transferRowHeaders(wsformat)
        self.transferRowHeaders(wsmultiple)
        self.transferRowHeaders(wssingle)
        self.transferRowHeaders(wsnothing)
        self.transferRowHeaders(wsopen)
        self.transferSheet1Rows()
        
    def incColIndex(self):
        self.colIndex += 1

    def incRowIndex(self):
        self.rowIndex += 1

    def incSheet2RowIndex(self):
        self.sheetFormatRowIndex += 1

    def getWorkbook(self):
        return self.workbook
    
    def mapColIndices(self):
        """Maps each header column to their respective list index"""
        headerIndex = self.getHeaderRowIndex()
        for currentIndex in range(1, self.MAX_COL+1):
            if (wsraw.cell(row=headerIndex, column=currentIndex).value is None):
                continue
            # pprint("current index is", self.headerIndexMap[wsraw.cell(row=headerIndex, column=currentIndex).value])
            self.headerIndexMap[wsraw.cell(row=headerIndex, column=currentIndex).value] = currentIndex - 2
        return
        
    def getHeaderRowIndex(self):
        """Returns the row index of the header row"""
        counter = 0
        for colCell in self.sheetRaw['A']:
            counter += 1
            if colCell.value == "Item":
                return counter
        
    def getColIndex(self, colName, headerRow, MAX_COL):
        """Returns the index of the column based on the given header column name"""
        for headerColIndex in range(1, MAX_COL):
            if self.sheetRaw.cell(row=headerRow, column=headerColIndex).value == colName:
                return headerColIndex
            
    def getColValues(self, targetCol):
        """Returns a list containing a tuple of all values from the respective column"""
        valueList = []
        for column in self.sheetRaw.iter_cols():
            colName = column[self.HEADER_ROW].value
            if colName == targetCol:
                for cell in column:
                    if cell.row > self.HEADER_ROW:
                        valueList.append(cell.value)
        return valueList

    def transferCol(self, colList, headerName):
        """Takes a header name and transfers that respective column over to Sheet2"""
        rowIndex = 1
        for colCellValue in colList:
            currentCell = self.sheetFormat.cell(row=self.sheetFormatRowIndex, column=self.colIndex)
            currentCell.value = colCellValue
            self.cellFormat(currentCell, headerName)
            self.incSheet2RowIndex()
        self.incColIndex()
        return
    
    def cellFormat(self, cell, headerName):
        """Formats the cell based on the headerName provided"""
        if headerName == "DATE":
            cell.number_format = "mm/dd/yyyy"
        if headerName == "AMOUNT":
            cell.number_format = "$#,##0.00"

    def transferSheet1Cols(self):
        """Copies to Sheet2 from Sheet1 each column respective to the headers from headerOrder"""
        for header in self.headerOrder:
            currentCol = self.getColValues(self.headerDict[header]["name"])
            self.transferCol(currentCol, header)
        return
    
    def getDonorFromValues(self):
        """Creates a donor object based on the given row values"""
        while self.rowIndex < self.MAX_ROW:
            row = self.sheetRaw[self.rowIndex]
            # print(row[0].value)
            self.incRowIndex()
        return

    def transferSheet1Rows(self):
        """Iterates through each row line item after the header in Sheet1 creates a Donor object
        Also writes the Donor object as a new line into Sheet2"""
        # use row[colindex].value to get the value

        for row in self.sheetRaw.iter_rows(min_row=self.HEADER_ROW+2, max_row=self.MAX_ROW):
            if not row[1].value:
                continue
            
            cellValues = []
            for colindex in range(1, self.MAX_COL):
                cellValues.append(row[colindex].value)

            newDonor = Donor(cellValues, self.headerIndexMap, self.headerDict, self.headerOrder)
            # call method to transfer Donor object into new line on sheet2
            self.transferDonor(newDonor)

            # if cellValues is not None:
            #     pprint(cellValues)
        return
    
    def transferDonor(self, donor):
        """Takes a donor object and writes each property value as a new line on Sheet2"""
        colIndex = 1
        for header in self.headerOrder:
            currentCell = self.sheetFormat.cell(row=self.sheetFormatRowIndex, column=colIndex)
            currentDonorValue = donor.properties[header]
            currentCell.value = currentDonorValue
            self.cellFormat(currentCell, header)
            colIndex += 1
        self.incSheet2RowIndex()

    def transferRowHeaders(self, worksheet):
        """Copies to Sheet2 each header respective to the headers from headerOrder"""
        colindex, rowindex = 1, 1
        for header in self.headerOrder:
            worksheet.cell(row=rowindex, column=colindex, value=self.headerDict[header]["name"])
            colindex += 1
        return
    
path_to_xlsx = path.abspath(path.join(path.dirname(__file__), 'newack.xlsx'))
wb = openpyxl.load_workbook('newack.xlsx')

wsraw = wb['Sheet1']
wsformat = wb.create_sheet("Formatted")
wsmultiple = wb.create_sheet("Multiple")
wssingle = wb.create_sheet("Single")
wsnothing = wb.create_sheet("Nothing")
wsopen = wb.create_sheet("Open")

newReport = RevenueReport(wb, wsraw, wsformat, wsmultiple, wssingle, wsnothing, wsopen)
# newReport.mapColIndices()

# print(newReport.headerIndexMap)
# newReport.transferRowHeaders()

# newReport.transferRowHeaders(wsformat)
# newReport.transferRowHeaders(wsmultiple)
# newReport.transferRowHeaders(wssingle)
# newReport.transferRowHeaders(wsnothing)
# newReport.transferRowHeaders(wsopen)
# newReport.transferSheet1Rows()


# test for getting the row headers

# for x in range(1, len(newReport.headerOrder)+1):
#     print(ws2.cell(row=1,column=x).value)


fixedBook = newReport.getWorkbook()
fixedBook.save("superAck.xlsx")
